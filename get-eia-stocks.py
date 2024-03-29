#!/usr/bin/env python
from dateutil.parser import parse
from datetime import datetime
import pandas as pd
import re


# TODO: move this function to common.timezone as part of the parser
def remove_pattern(string, pattern):
    match = re.findall(pattern, string)
    new = string
    while match:
        new = new.replace(match.pop(), "")
    return new

class inventory_table_has_changed_from(object):
    """An expectation for checking that the inventory table has changed.

    Parameters
    ----------
    locator: tuple
        value pair like (By.ID, "id_value") used to find the table
    current: <type: table>
        previus version of the table to compare to
    """
    def __init__(self, locator, current):
        self.locator = locator
        self.current_length = len(current.find_elements_by_css_selector("tbody tr"))

    def __call__(self, browser):
        new_table = browser.find_element(*self.locator)
        new_length = len(new_table.find_elements_by_css_selector("tbody tr"))
        if self.current_length < new_length:
            return new_table
        else:
            return False

def parse_args():
    import argparse, sys, pytz

    parser = argparse.ArgumentParser()

    parser.add_argument(
        "from_date",
        metavar="from-date",
        help="Starting date in dataset"
    )
    parser.add_argument(
        "to_date",
        metavar="to-date",
        help="Closing date in dataset. Defaults to now in the given timezone"
    )
    parser.add_argument(
        "timezone",
        help="Timezone of 'from-date' & 'to-date' also alignment for the dataset"
    )
    parser.add_argument(
        "datetime_format",
        metavar="datetime-format",
        help="The datetime format of the dataset, example: RFC3339, UNIX, JSON, or any datetime format string.",
    )
    parser.add_argument(
        "--save-to", "-s",
        help="Stores the dataset in the given filename. Supported extensions are .csv (default if extension is missing) or .xlsx file. If .xlsx and a second argument (comma-separated) value is given, it is taken to be the name of the sheet."
    )
    # TODO: add interactive option
    # TODO: add verbose option

    # TODO: fix help string in save-to argument
    # TODO: fix help string in to-date
    args = parser.parse_args()

    if args.to_date.lower() == "now":
        args.to_date = datetime.now(tz=pytz.timezone(args.timezone)).isoformat()

    return args._get_args(), dict(args._get_kwargs())

# TODO: implement parse "better" and "worse" than expected
def request_data(*args, **kwargs):
    import locale
    from selenium import webdriver
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from pyCBT.common.timezone import parse_tz, timezone_shift

    locale.setlocale(locale.LC_TIME, "en_US")
    from_date = parse_tz(kwargs.get("from_date"), in_tz=kwargs.get("timezone"))
    to_date = parse_tz(kwargs.get("to_date"), in_tz=kwargs.get("timezone"))

    browser = webdriver.Chrome()
    browser.get("https://www.investing.com/economic-calendar/eia-crude-oil-inventories-75")

    inv_table = browser.find_element(By.ID, "eventHistoryTable75")
    last_date_str = inv_table.find_element_by_css_selector("tbody tr:last-child td").text
    last_record_date = parse_tz(remove_pattern(last_date_str, r"\(\w+\)"), in_tz="America/New_York")

    wait = WebDriverWait(browser, 10)
    while last_record_date > from_date:
        show_more = wait.until(EC.element_to_be_clickable((By.ID, "showMoreHistory75")))
        browser.execute_script("arguments[0].click();", show_more)

        inv_table = wait.until(inventory_table_has_changed_from((By.ID, "eventHistoryTable75"), inv_table))
        last_date_str = inv_table.find_element_by_css_selector("tbody tr:last-child td").text
        last_record_date = parse_tz(remove_pattern(last_date_str, r"\(\w+\)"), in_tz="America/New_York")

    table = pd.read_html(u"<table>"+inv_table.get_attribute("innerHTML")+u"</table>")[0]
    table.insert(0, "Datetime", value=table["Release Date"]+" "+table["Time"])
    better = map(lambda span: "better" in span.get_attribute("title").lower() if span.get_attribute("title").strip() else None, inv_table.find_elements_by_tag_name("span"))
    table.insert(table.columns.size, "Better", value=better)
    table["Datetime"] = table["Datetime"].apply(remove_pattern, args=(r"\(\w+\)",))
    table["Datetime"] = table["Datetime"].apply(
        timezone_shift,
        args=("America/New_York", kwargs.get("timezone"), kwargs.get("datetime_format"))
    )
    mask = [not (from_date <= parse_tz(release_date, in_tz="America/New_York") <= to_date) for release_date in table["Datetime"]]
    table.drop(table.index[mask], axis="index", inplace=True)
    table.drop(["Release Date", "Time", "Unnamed: 5"], axis="columns", inplace=True)
    table.set_index("Datetime", inplace=True)
    table = table.applymap(lambda cell: eval(cell.strip("M")) if type(cell) == str else cell)

    locale.resetlocale(locale.LC_TIME)
    return table

def dump_data(*args, **kwargs):
    import os, sys, string
    import pandas as pd
    from pyCBT.common.files import exist
    from openpyxl import load_workbook

    dataframe, = args
    if kwargs.get("save_to") is not None:
        if "," in kwargs["save_to"]: filename, sheetname = kwargs["save_to"].split(",")
        else: filename, sheetname = kwargs.get("save_to"), "sheet_001"
        if filename.endswith(".xlsx") and exist(filename):
            # TODO: if the file exist, ask the user
            book = load_workbook(filename)
            with pd.ExcelWriter(filename, engine="openpyxl") as excel_writer:
                excel_writer.book = book
                if sheetname in book.sheetnames:
                    match = re.match("(\w+)_(\d+)", sheetname)
                    if not match: sheetname += "_{0:03d}"
                    else: sheetname = string.join([match.groups()[0], "{0:03d}"], "_")
                i = 1
                while sheetname.format(i) in book.sheetnames: i += 1
                sheetname = sheetname.format(i)
                dataframe.to_excel(excel_writer, sheet_name=sheetname)
                excel_writer.save()
            book.close()
        elif filename.endswith(".xlsx"):
            with pd.ExcelWriter(filename, engine="openpyxl") as excel_writer:
                dataframe.to_excel(excel_writer, sheet_name=sheetname)
                excel_writer.save()
        else:
            # TODO: if the file exist, ask the user
            if filename.split(".")[-1] != "csv": filename += ".csv"
            dataframe.reset_index().to_csv(filename, index=False, line_terminator=os.linesep)
    else:
        dataframe.reset_index().to_csv(sys.stdout, index=False, line_terminator=os.linesep)

    return None

if __name__ == "__main__":
    args, kwargs = parse_args()
    table = request_data(*args, **kwargs)
    dump_data(table, **kwargs)
